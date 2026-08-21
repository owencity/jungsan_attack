#!/usr/bin/env python3
"""
Liquibase changelog(YAML) → 테이블 명세 엑셀(5시트) 생성기.

00-README.md의 원래 계획은 "실행 중인 DB의 information_schema·pg_description에서
자동 생성"이었다. 지금은 Docker가 꺼져 있어 그 방식을 못 쓴다 — 대신 changelog
YAML 자체를 단일 진실 소스로 파싱한다. changelog가 바뀌면 이 스크립트를 다시
돌리기만 하면 되므로, "손으로 관리하는 명세서는 반드시 어긋난다"는 원칙은
여전히 지켜진다(파싱 대상이 DB냐 YAML이냐의 차이일 뿐).

사용법: python scripts/generate_table_spec.py
출력:   docs/table-spec.xlsx
"""
import yaml
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CHANGELOG_DIR = ROOT / "server" / "src" / "main" / "resources" / "db" / "changelog"
MASTER = CHANGELOG_DIR / "db.changelog-master.yaml"
OUT = ROOT / "docs" / "table-spec.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2C74D6")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_yaml(path):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def included_files():
    doc = load_yaml(MASTER)
    files = []
    for item in doc["databaseChangeLog"]:
        if "include" in item:
            files.append(ROOT / "server" / "src" / "main" / "resources" / item["include"]["file"])
    return files


def parse_changelogs():
    """모든 changeSet을 훑어 tables / columns / constraints(인덱스류) 를 뽑는다."""
    tables = {}      # name -> {remarks, source_file, source_changeset}
    columns = []      # list of dict
    constraints = []  # list of dict (PK/UNIQUE/INDEX/FK)

    for path in included_files():
        doc = load_yaml(path)
        for entry in doc["databaseChangeLog"]:
            cs = entry.get("changeSet")
            if not cs:
                continue
            cs_id = cs["id"]
            for change in cs.get("changes", []):
                if "createTable" in change:
                    ct = change["createTable"]
                    tname = ct["tableName"]
                    tables[tname] = {
                        "remarks": ct.get("remarks", ""),
                        "source_file": path.name,
                        "source_changeset": cs_id,
                    }
                    for col in ct.get("columns", []):
                        c = col["column"]
                        constr = c.get("constraints", {}) or {}
                        fk = ""
                        if constr.get("foreignKeyName"):
                            fk = f'{constr.get("references", "")} ({constr["foreignKeyName"]})'
                        columns.append({
                            "table": tname,
                            "column": c["name"],
                            "type": c.get("type", ""),
                            "nullable": "N" if constr.get("nullable") is False else "Y",
                            "pk": "PK" if constr.get("primaryKey") else "",
                            "default": c.get("defaultValue", c.get("defaultValueNumeric", c.get("defaultValueBoolean", ""))),
                            "fk": fk,
                            "remarks": c.get("remarks", ""),
                        })

                elif "addUniqueConstraint" in change:
                    u = change["addUniqueConstraint"]
                    constraints.append({
                        "table": u["tableName"], "name": u["constraintName"],
                        "type": "UNIQUE", "columns": u["columnNames"],
                        "remarks": u.get("remarks", ""),
                    })
                elif "addPrimaryKey" in change:
                    p = change["addPrimaryKey"]
                    constraints.append({
                        "table": p["tableName"], "name": p["constraintName"],
                        "type": "PRIMARY KEY", "columns": p["columnNames"],
                        "remarks": p.get("remarks", ""),
                    })
                elif "createIndex" in change:
                    ix = change["createIndex"]
                    cols = ", ".join(c["column"]["name"] for c in ix["columns"])
                    constraints.append({
                        "table": ix["tableName"], "name": ix["indexName"],
                        "type": "INDEX", "columns": cols,
                        "remarks": ix.get("remarks", ""),
                    })
                elif "addForeignKeyConstraint" in change:
                    fk = change["addForeignKeyConstraint"]
                    constraints.append({
                        "table": fk["baseTableName"], "name": fk["constraintName"],
                        "type": "FOREIGN KEY",
                        "columns": f'{fk["baseColumnNames"]} → {fk["referencedTableName"]}.{fk["referencedColumnNames"]}',
                        "remarks": "순환 참조 — 뒤늦게 추가된 FK (참고: ERD.md §3)",
                    })
    return tables, columns, constraints


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    tables, columns, constraints = parse_changelogs()
    wb = Workbook()

    # ── 1. 스키마 (개요) ──────────────────────────
    ws = wb.active
    ws.title = "스키마"
    ws.append(["항목", "값"])
    rows = [
        ("DB 엔진", "MySQL 8.4 LTS (ADR-011)"),
        ("변경 관리", "Liquibase (YAML)"),
        ("소스 경로", "server/src/main/resources/db/changelog/"),
        ("changelog 파일 수", len(included_files())),
        ("테이블 수", len(tables)),
        ("컬럼 수", len(columns)),
        ("제약·인덱스 수", len(constraints)),
        ("생성 방식", "changelog YAML 파싱 (scripts/generate_table_spec.py) — 실행 중 DB 대상 아님"),
        ("실행 검증", "⚠ 미완료 — Docker 켜고 liquibase update 로 확인할 것"),
        ("관련 ADR", "001(유리수) · 002(아웃박스) · 005(미저장) · 008(동시성) · 009(Group) · 011(MySQL) · 012(오프로드)"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 2)
    autofit(ws, [22, 70])

    # ── 2. 테이블목록 ──────────────────────────
    ws = wb.create_sheet("테이블목록")
    ws.append(["테이블명", "설명(remarks)", "컬럼 수", "정의 파일", "changeSet id"])
    for name in sorted(tables):
        t = tables[name]
        col_count = sum(1 for c in columns if c["table"] == name)
        ws.append([name, t["remarks"], col_count, t["source_file"], t["source_changeset"]])
    style_header(ws, 5)
    autofit(ws, [22, 60, 10, 24, 26])

    # ── 3. 컬럼목록 ──────────────────────────
    ws = wb.create_sheet("컬럼목록")
    ws.append(["테이블", "컬럼", "타입", "NULL 허용", "PK", "기본값", "FK 참조", "설명(remarks)"])
    for c in columns:
        ws.append([c["table"], c["column"], c["type"], c["nullable"], c["pk"],
                   str(c["default"]) if c["default"] not in (None, "") else "", c["fk"], c["remarks"]])
    style_header(ws, 8)
    autofit(ws, [18, 22, 14, 10, 6, 12, 34, 70])

    # ── 4. 코멘트 (컬럼 remarks만 모은 요약) ──────
    ws = wb.create_sheet("코멘트")
    ws.append(["테이블", "컬럼", "설명(remarks)"])
    missing = 0
    for c in columns:
        ws.append([c["table"], c["column"], c["remarks"]])
        if not c["remarks"]:
            missing += 1
    style_header(ws, 3)
    autofit(ws, [22, 22, 90])
    if missing:
        ws.append([])
        ws.append([f"⚠ remarks 누락 {missing}건 — 모든 컬럼에 주석이 있어야 한다는 원칙 위반"])

    # ── 5. 인덱스 (PK·UNIQUE·INDEX·FK) ──────────
    ws = wb.create_sheet("인덱스")
    ws.append(["테이블", "제약/인덱스명", "종류", "대상 컬럼", "설명(remarks)"])
    for c in constraints:
        cols = c["columns"] if isinstance(c["columns"], str) else ", ".join(c["columns"])
        ws.append([c["table"], c["name"], c["type"], cols, c["remarks"]])
    style_header(ws, 5)
    autofit(ws, [22, 34, 14, 40, 60])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"생성 완료: {OUT}")
    print(f"  테이블 {len(tables)}개 · 컬럼 {len(columns)}개 · 제약/인덱스 {len(constraints)}개")
    if missing:
        print(f"  ⚠ remarks 누락 {missing}건")


if __name__ == "__main__":
    main()
